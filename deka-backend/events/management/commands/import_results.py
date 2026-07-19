import os
from datetime import timedelta
from pathlib import Path
import re

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from events.models import AthleteResult, DekaEvent


class Command(BaseCommand):
  help = "Import athlete results from an .xlsx file into the database"
  NAME_COLUMN = 0
  CATEGORY_COLUMN = 1
  GENDER_COLUMN = 2
  DEKA_TYPE_COLUMN = 3
  TOTAL_TIME_COLUMN = 4
  RUN_TIME_COLUMNS = list({5, 7, 9, 11, 13, 15, 17, 19, 21, 23})
  ZONE_TIME_COLUMNS = list({6, 8, 10, 12, 14, 16, 18, 20, 22, 24})

  def add_arguments(self, parser):
    parser.add_argument("file_path", type=str, help="Path to the .xlsx file to import")
    parser.add_argument(
      "--event",
      type=str,
      default=None,
      help="Event name to associate with all imported rows. Defaults to the first sheet name.",
    )

  def import_sheet(self, sheet, event):
    print(f"Importing from sheet '{sheet.title}' with {sheet.max_row - 1} data rows")
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
      raise CommandError("The selected sheet is empty")

    headers = [self._normalize_header(value) for value in rows[0]]
    data_rows = rows[1:]

    if not headers:
      raise CommandError("No headers were found in the selected sheet")

    imported = 0
    skipped = 0

    with transaction.atomic():
      for index, row in enumerate(data_rows, start=2):
        if (len(row) < 24):
          skipped += 1
          print(f"Row {index} skipped: Not enough columns (found {len(row)}, expected at least 24)")
          continue

        athlete_name = row[self.NAME_COLUMN]
        if not athlete_name:
          skipped += 1
          print(f"Row {index} skipped: Athlete name is missing")
          continue

        deka_type = row[self.DEKA_TYPE_COLUMN]
        if not deka_type:
          skipped += 1
          print(f"Row {index} skipped: DEKA type is missing")
          continue

        orig_category = row[self.CATEGORY_COLUMN]
        if not orig_category:
          skipped += 1
          print(f"Row {index} skipped: Category is missing")
          continue

        age_group = ""
        gender = row[self.GENDER_COLUMN]
        total_time_str = row[self.TOTAL_TIME_COLUMN]
        total_time = self._parse_duration(total_time_str)

        run_times = [self._parse_duration(row[i]) for i in self.RUN_TIME_COLUMNS]
        zone_times = [self._parse_duration(row[i]) for i in self.ZONE_TIME_COLUMNS]

        # Parse the category to understand gender and age group
        if not gender:
          if "Men" in orig_category or "Male" in orig_category:
            gender = "Male"
          elif "Women" in orig_category or "Female" in orig_category:
            gender = "Female"
          elif "Mixed" in orig_category or "co-ed" in orig_category.lower():
            gender = "Mixed"

        category = orig_category
        if "Elite" in orig_category:
          category = "Elite"
        elif "Age Group" in orig_category or "AgeGroup" in orig_category:
          category = "Age Group"
          match = re.search(r"\d{2}-\d{2}", orig_category)
          if match:
            age_group = match.group(0)
        elif "Open" in orig_category:
          category = "Open"

        athlete_result, created = AthleteResult.objects.get_or_create(
          event=event,
          athlete_name=str(athlete_name).strip(),
          deka_type=str(deka_type).strip(),
          category=str(category or "").strip() or "",
          defaults={
            "gender": gender or "",
            "age_group": age_group,
            "total_time": total_time,
            "run_1": run_times[0],
            "run_2": run_times[1],
            "run_3": run_times[2],
            "run_4": run_times[3],
            "run_5": run_times[4],
            "run_6": run_times[5],
            "run_7": run_times[6],
            "run_8": run_times[7],
            "run_9": run_times[8],
            "run_10": run_times[9],
            "zone_1": zone_times[0],
            "zone_2": zone_times[1],
            "zone_3": zone_times[2],
            "zone_4": zone_times[3],
            "zone_5": zone_times[4],
            "zone_6": zone_times[5],
            "zone_7": zone_times[6],
            "zone_8": zone_times[7],
            "zone_9": zone_times[8],
            "zone_10": zone_times[9],
          },
        )

        if created:
          imported += 1
          #print(f"Row {index} imported: {athlete_result}")
        else:
          # Update the existing record
          athlete_result.gender = gender
          athlete_result.age_group = age_group
          athlete_result.total_time = total_time
          athlete_result.run_1 = run_times[0]
          athlete_result.run_2 = run_times[1]
          athlete_result.run_3 = run_times[2]
          athlete_result.run_4 = run_times[3]
          athlete_result.run_5 = run_times[4]
          athlete_result.run_6 = run_times[5]
          athlete_result.run_7 = run_times[6]
          athlete_result.run_8 = run_times[7]
          athlete_result.run_9 = run_times[8]
          athlete_result.run_10 = run_times[9]
          athlete_result.zone_1 = zone_times[0]
          athlete_result.zone_2 = zone_times[1]
          athlete_result.zone_3 = zone_times[2]
          athlete_result.zone_4 = zone_times[3]
          athlete_result.zone_5 = zone_times[4]
          athlete_result.zone_6 = zone_times[5]
          athlete_result.zone_7 = zone_times[6]
          athlete_result.zone_8 = zone_times[7]
          athlete_result.zone_9 = zone_times[8]
          athlete_result.zone_10 = zone_times[9]
          athlete_result.save()

  def handle(self, *args, **options):
    # Get the input file from the command line arguments and validate it
    input_path = Path(options["file_path"]).expanduser().resolve()
    if not input_path.exists():
        raise CommandError(f"File not found: {input_path}")
    if input_path.suffix.lower() != ".xlsx":
        raise CommandError("Only .xlsx files are supported")

    # Get the event name from the command line argument and create or retrieve the event from the database
    event_name = options["event"] or input_path.stem
    event, created = DekaEvent.objects.get_or_create(name=event_name)
    if created:
        self.stdout.write(self.style.SUCCESS(f"Created event: {event.name}"))
    else:
        self.stdout.write(f"Using existing event: {event.name}")

    # Load the workbook
    workbook = openpyxl.load_workbook(input_path, data_only=True)

    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      self.import_sheet(sheet, event)


  def _normalize_header(self, value):
    if value is None:
      return ""
    return str(value).strip().lower().replace(" ", "_").replace("-", "_")

  def _get_value(self, values, possible_keys):
    for key in possible_keys:
      if key in values and values[key] not in (None, ""):
        return values[key]
    return None

  def _parse_duration(self, value):
    if value in (None, ""):
      return None
    if isinstance(value, timedelta):
      return value
    if isinstance(value, (int, float)):
      return timedelta(seconds=float(value))
    text = str(value).strip()
    if not text:
      return None
    if ":" in text:
      parts = text.split(":")
      if len(parts) == 2:
        minutes, seconds = parts
        return timedelta(minutes=int(minutes), seconds=float(seconds))
      if len(parts) == 3:
        hours, minutes, seconds = parts
        return timedelta(hours=int(hours), minutes=int(minutes), seconds=float(seconds))
    try:
      seconds = float(text)
      return timedelta(seconds=seconds)
    except ValueError:
      return None
