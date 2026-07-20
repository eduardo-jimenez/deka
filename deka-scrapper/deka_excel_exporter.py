from openpyxl import Workbook
from results_data import DekaResults, AthleteResult, DekaType, DekaGender



def get_athlete_row(athlete:AthleteResult, deka_type:str, category_name:str) -> list:
  gender = DekaGender.get_name(athlete.gender)
  row = [athlete.name, deka_type, gender, category_name, athlete.time]
  for i in range(0, len(athlete.run_times)):
    row.append(athlete.run_times[i])
    row.append(athlete.zone_times[i])

  return row

def get_athlete_title_row() -> list:
  title = [
    "Name", "DEKA Type", "Gender", "Category", "Mark", 
    "Run 1", "Zone 1",
    "Run 2", "Zone 2",
    "Run 3", "Zone 3",
    "Run 4", "Zone 4",
    "Run 5", "Zone 5",
    "Run 6", "Zone 6",
    "Run 7", "Zone 7",
    "Run 8", "Zone 8",
    "Run 9", "Zone 9",
    "Run 10", "Zone 10"
  ]

  return title

def export_to_excel(deka:DekaResults, filename:str):
  print(f"Exporting {deka} to {filename} (total of {len(deka.types)} types)")

  wb = Workbook()
  ws = wb.active
  if ws is not None:
    wb.remove(ws)

  for deka_type in deka.types:
    #print(f"Adding sheet for {deka_type}")
    deka_type_name = DekaType.get_name(deka_type.type)

    ws = wb.create_sheet(title=deka_type.name[:31])
    ws.append(get_athlete_title_row())

    for category in deka_type.categories:
      #print(f"Category {category.name} has {len(category.athletes)} athletes")

      for athlete in category.athletes:
        try:
          row = get_athlete_row(athlete, deka_type_name, category.name)
          ws.append(row)
        except Exception as e:
          print(f"Error extracting and adding row for athlete {athlete.name} in {category.name}")
  
  # add a sheet with the event info
  ws = wb.create_sheet(title="Event Info")
  ws.append(deka.name)
  ws.append(deka.url)
  ws.append(deka.city)
  ws.append(deka.date)

  # save to file  
  wb.save(filename)

